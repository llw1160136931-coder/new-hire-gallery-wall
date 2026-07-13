import re
from dataclasses import dataclass
from pathlib import Path

from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from api.models import Profile


HEADER_ALIASES = {
    'name': ('姓名', '名字', '管理员姓名'),
    'username': ('账号', '用户名', '登录账号'),
    'password': ('密码', '初始密码'),
    'role': ('角色', '身份'),
    'workplace': ('工作单位', '单位', '公司'),
}
REQUIRED_FIELDS = ('name', 'username', 'password', 'role')
ADMIN_ROLE_VALUES = {'管理员', 'admin'}
ZERO_NUMBER_FORMAT = re.compile(r'^0+$')


@dataclass(frozen=True)
class AdminRow:
    row_number: int
    name: str
    username: str
    password: str
    workplace: str
    role: str


class Command(BaseCommand):
    help = '从 Excel 安全批量导入系统管理员，密码使用 Django 哈希保存。'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', help='管理员 Excel 文件路径（仅支持 .xlsx）')
        parser.add_argument('--sheet', help='工作表名称；不填写时使用第一个工作表')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='只检查文件和数据，不写入数据库',
        )
        parser.add_argument(
            '--update-existing',
            action='store_true',
            help='更新已存在的非超级管理员账号；默认遇到重复账号就停止',
        )

    def handle(self, *args, **options):
        excel_path = Path(options['excel_path']).expanduser().resolve()
        if excel_path.suffix.lower() != '.xlsx':
            raise CommandError('只支持 .xlsx 文件。')
        if not excel_path.is_file():
            raise CommandError('找不到管理员 Excel 文件。')

        try:
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError('管理员 Excel 文件无法读取，请确认文件没有损坏。') from exc

        try:
            worksheet = self._select_worksheet(workbook, options.get('sheet'))
            rows = self._read_rows(worksheet)
        finally:
            workbook.close()

        existing_users = {
            user.username: user
            for user in User.objects.select_related('profile').filter(
                username__in=[row.username for row in rows]
            )
        }
        errors = self._validate_rows(rows, existing_users, options['update_existing'])
        if errors:
            formatted = '\n'.join(f'- 第 {row_number} 行：{message}' for row_number, message in errors)
            raise CommandError(f'管理员导入检查失败，共 {len(errors)} 个问题：\n{formatted}')

        new_count = sum(row.username not in existing_users for row in rows)
        update_count = len(rows) - new_count
        if options['dry_run']:
            self.stdout.write(
                self.style.SUCCESS(
                    f'检查通过：共 {len(rows)} 名管理员，可新增 {new_count} 名，可更新 {update_count} 名；数据库未改动。'
                )
            )
            return

        with transaction.atomic():
            for row in rows:
                self._save_admin(row, existing_users.get(row.username))

        self.stdout.write(
            self.style.SUCCESS(
                f'管理员导入完成：新增 {new_count} 名，更新 {update_count} 名。'
            )
        )
        self.stdout.write('安全提示：请立即删除服务器上的管理员 Excel 文件。')

    def _select_worksheet(self, workbook, sheet_name):
        if not sheet_name:
            return workbook.worksheets[0]
        if sheet_name not in workbook.sheetnames:
            raise CommandError('找不到指定工作表。')
        return workbook[sheet_name]

    def _read_rows(self, worksheet):
        row_iterator = worksheet.iter_rows()
        try:
            header_cells = next(row_iterator)
        except StopIteration as exc:
            raise CommandError('管理员 Excel 文件没有数据。') from exc

        headers = [self._cell_text(cell) for cell in header_cells]
        indexes = self._map_headers(headers)
        rows = []
        for row_number, cells in enumerate(row_iterator, start=2):
            if not any(self._cell_text(cell) for cell in cells):
                continue
            rows.append(AdminRow(
                row_number=row_number,
                name=self._value_at(cells, indexes['name']),
                username=self._value_at(cells, indexes['username']),
                password=self._value_at(cells, indexes['password']),
                workplace=self._value_at(cells, indexes.get('workplace')),
                role=self._value_at(cells, indexes['role']).lower(),
            ))
        if not rows:
            raise CommandError('管理员 Excel 文件没有可导入的数据。')
        return rows

    def _map_headers(self, headers):
        indexes = {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if alias in headers:
                    indexes[field] = headers.index(alias)
                    break
        missing = [HEADER_ALIASES[field][0] for field in REQUIRED_FIELDS if field not in indexes]
        if missing:
            raise CommandError(f'管理员 Excel 缺少必填列：{"、".join(missing)}。')
        return indexes

    def _validate_rows(self, rows, existing_users, update_existing):
        errors = []
        seen_usernames = set()
        for row in rows:
            if not row.name:
                errors.append((row.row_number, '姓名不能为空。'))
            if not row.username:
                errors.append((row.row_number, '账号不能为空。'))
            elif len(row.username) > User._meta.get_field('username').max_length:
                errors.append((row.row_number, '账号长度超过系统限制。'))
            elif row.username in seen_usernames:
                errors.append((row.row_number, 'Excel 内存在重复管理员账号。'))
            seen_usernames.add(row.username)

            if row.role not in ADMIN_ROLE_VALUES:
                errors.append((row.row_number, '角色必须填写为“管理员”。'))
            if not row.password:
                errors.append((row.row_number, '密码不能为空。'))
            else:
                try:
                    validate_password(row.password, user=User(username=row.username))
                except ValidationError as exc:
                    errors.append((row.row_number, '管理员密码不符合安全要求：' + '；'.join(exc.messages)))

            existing_user = existing_users.get(row.username)
            if existing_user and not update_existing:
                errors.append((row.row_number, '账号已存在；如确认更新，请使用 --update-existing。'))
            if existing_user and existing_user.is_superuser:
                errors.append((row.row_number, '超级管理员账号禁止通过批量导入命令修改。'))
        return errors

    def _save_admin(self, row, existing_user):
        user = existing_user or User(username=row.username)
        user.is_active = True
        user.is_staff = True
        user.is_superuser = False
        user.set_password(row.password)
        user.save()
        profile, _ = Profile.objects.get_or_create(
            user=user,
            defaults={'name': row.name, 'role': Profile.Role.ADMIN},
        )
        profile.name = row.name
        profile.role = Profile.Role.ADMIN
        if row.workplace:
            profile.workplace = row.workplace
        profile.save()

    def _value_at(self, cells, index):
        if index is None or index >= len(cells):
            return ''
        return self._cell_text(cells[index])

    def _cell_text(self, cell):
        value = cell.value
        if value is None:
            return ''
        if isinstance(value, bool):
            return '是' if value else '否'
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            number_format = str(cell.number_format or '')
            if float(value).is_integer():
                integer = int(value)
                if ZERO_NUMBER_FORMAT.fullmatch(number_format):
                    return f'{integer:0{len(number_format)}d}'
                return str(integer)
        return str(value).strip()
