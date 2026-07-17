import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from api.models import Profile


REQUIRED_HEADERS = {
    'name': ('姓名', '名字', '学员姓名'),
    'mbti': ('MBTI类型', 'MBTI'),
    'training_group': ('小组', '分组', '组别'),
}
SUPPORTED_HEADERS = {
    alias
    for aliases in REQUIRED_HEADERS.values()
    for alias in aliases
}
MBTI_PATTERN = re.compile(r'^\s*([A-Za-z]{4})(?:\s*[（(].*[）)])?\s*$')
GROUP_PATTERN = re.compile(r'^\s*(?:第\s*)?([1-6])(?:\s*组)?\s*$')


@dataclass(frozen=True)
class StudentProfileRow:
    row_number: int
    name: str
    mbti: str
    training_group: str


class Command(BaseCommand):
    help = '按姓名严格匹配已有学员，批量更新 MBTI 和小组，不修改账号或密码。'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', help='Excel 文件路径（仅支持 .xlsx）')
        parser.add_argument('--sheet', help='工作表名称；不填写时使用第一个工作表')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='只检查姓名匹配和资料格式，不写入数据库',
        )

    def handle(self, *args, **options):
        excel_path = Path(options['excel_path']).expanduser().resolve()
        if excel_path.suffix.lower() != '.xlsx':
            raise CommandError('只支持 .xlsx 文件。')
        if not excel_path.is_file():
            raise CommandError('找不到 Excel 文件。')

        try:
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError('Excel 文件无法读取，请确认文件没有损坏。') from exc

        try:
            worksheet = self._select_worksheet(workbook, options.get('sheet'))
            rows, ignored_headers = self._read_rows(worksheet)
        finally:
            workbook.close()

        profiles_by_name = defaultdict(list)
        profiles = Profile.objects.select_related('user').filter(
            name__in={row.name for row in rows}
        )
        for profile in profiles:
            profiles_by_name[profile.name].append(profile)

        errors = self._validate_rows(rows, profiles_by_name)
        if errors:
            formatted = '\n'.join(
                f'- 第 {row_number} 行：{message}'
                for row_number, message in errors
            )
            raise CommandError(f'学员资料导入检查失败，共 {len(errors)} 个问题：\n{formatted}')

        if ignored_headers:
            self.stdout.write(
                self.style.WARNING(
                    f'以下列当前不会导入：{"、".join(ignored_headers)}'
                )
            )

        if options['dry_run']:
            self.stdout.write(
                self.style.SUCCESS(
                    f'检查通过：共 {len(rows)} 名学员，姓名均唯一匹配；数据库未改动。'
                )
            )
            return

        with transaction.atomic():
            for row in rows:
                profile = profiles_by_name[row.name][0]
                profile.mbti = self._normalize_mbti(row.mbti)
                profile.training_group = self._normalize_group(row.training_group)
                profile.save(update_fields=['mbti', 'training_group', 'updated_at'])

        self.stdout.write(
            self.style.SUCCESS(
                f'学员资料更新完成：已更新 {len(rows)} 名学员的 MBTI 和小组，未修改账号或密码。'
            )
        )
        self.stdout.write('安全提示：请立即删除服务器上的原始 Excel 文件。')

    def _select_worksheet(self, workbook, sheet_name):
        if not sheet_name:
            return workbook.worksheets[0]
        if sheet_name not in workbook.sheetnames:
            raise CommandError('找不到指定工作表。')
        return workbook[sheet_name]

    def _read_rows(self, worksheet):
        row_iterator = worksheet.iter_rows()
        try:
            header_cells = next(row_iterator)
        except StopIteration as exc:
            raise CommandError('Excel 文件没有数据。') from exc

        headers = [self._cell_text(cell) for cell in header_cells]
        header_indexes = self._map_headers(headers)
        ignored_headers = [
            header for header in headers
            if header and header not in SUPPORTED_HEADERS
        ]
        rows = []
        for row_number, cells in enumerate(row_iterator, start=2):
            if not any(self._cell_text(cell) for cell in cells):
                continue
            rows.append(StudentProfileRow(
                row_number=row_number,
                name=self._value_at(cells, header_indexes['name']),
                mbti=self._value_at(cells, header_indexes['mbti']),
                training_group=self._value_at(cells, header_indexes['training_group']),
            ))

        if not rows:
            raise CommandError('Excel 文件没有可导入的学员资料。')
        return rows, ignored_headers

    def _map_headers(self, headers):
        indexes = {}
        for field, aliases in REQUIRED_HEADERS.items():
            for alias in aliases:
                if alias in headers:
                    indexes[field] = headers.index(alias)
                    break
        missing = [
            aliases[0]
            for field, aliases in REQUIRED_HEADERS.items()
            if field not in indexes
        ]
        if missing:
            raise CommandError(f'Excel 缺少必填列：{"、".join(missing)}。')
        return indexes

    def _validate_rows(self, rows, profiles_by_name):
        errors = []
        seen_names = set()
        valid_mbti_values = set(Profile.Mbti.values)
        for row in rows:
            if not row.name:
                errors.append((row.row_number, '姓名不能为空。'))
            elif row.name in seen_names:
                errors.append((row.row_number, 'Excel 内存在重复姓名。'))
            seen_names.add(row.name)

            normalized_mbti = self._normalize_mbti(row.mbti)
            if row.mbti and not normalized_mbti:
                errors.append((row.row_number, 'MBTI 格式不正确，应为 4 位类型代码，可带中文说明。'))
            elif normalized_mbti and normalized_mbti not in valid_mbti_values:
                errors.append((row.row_number, f'不支持的 MBTI 类型：{normalized_mbti}。'))

            if not self._normalize_group(row.training_group):
                errors.append((row.row_number, '小组只支持第1组至第6组，且不能留空。'))

            matches = profiles_by_name.get(row.name, [])
            if not row.name:
                continue
            if not matches:
                errors.append((row.row_number, '数据库中找不到该姓名的学员。'))
            elif len(matches) > 1:
                errors.append((row.row_number, '数据库中存在多个同名账号，为避免写错人已停止导入。'))
            elif self._is_admin(matches[0]):
                errors.append((row.row_number, '该姓名对应管理员账号，禁止通过学员资料命令修改。'))
        return errors

    @staticmethod
    def _is_admin(profile):
        return (
            profile.role == Profile.Role.ADMIN
            or profile.user.is_staff
            or profile.user.is_superuser
        )

    @staticmethod
    def _normalize_mbti(value):
        normalized = str(value or '').strip()
        if not normalized:
            return ''
        match = MBTI_PATTERN.fullmatch(normalized)
        if not match:
            return ''
        return match.group(1).upper()

    @staticmethod
    def _normalize_group(value):
        match = GROUP_PATTERN.fullmatch(str(value or '').strip())
        return match.group(1) if match else ''

    def _value_at(self, cells, index):
        if index is None or index >= len(cells):
            return ''
        return self._cell_text(cells[index])

    @staticmethod
    def _cell_text(cell):
        value = cell.value
        if value is None:
            return ''
        if isinstance(value, bool):
            return '是' if value else '否'
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if float(value).is_integer():
                return str(int(value))
        return str(value).strip()
